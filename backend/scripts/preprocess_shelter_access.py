"""쉼터·인구 원자료를 행정동 코드 기준의 실시간 점수 입력자료로 전처리한다."""

# 현용 점수체계 수정 26-08-06 13-00

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


DOT_NAME_REPLACEMENTS = {
    "신천1.2동": "신천1·2동",
    "불로.봉무동": "불로·봉무동",
    "내당2.3동": "내당2·3동",
    "비산2.3동": "비산2·3동",
    "수성2.3가동": "수성2·3가동",
    "두류1.2동": "두류1·2동",
}
BRANCH_OFFICE_TARGETS = {
    "논공읍공단출장소": "논공읍",
    "다사읍서재출장소": "다사읍",
}


def normalize_dong_name(name: object) -> str:
    """자료원마다 다른 점 표기와 출장소 명칭을 행정동 명칭으로 통일한다."""
    normalized = re.sub(r"\s+", "", str(name or "").strip())
    normalized = DOT_NAME_REPLACEMENTS.get(normalized, normalized)
    return BRANCH_OFFICE_TARGETS.get(normalized, normalized)


def read_canonical_regions(path: Path) -> dict[tuple[str, str], dict[str, str]]:
    """기존 지도와 같은 150개 행정동 코드 사전을 읽는다."""
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        rows = list(csv.DictReader(source))
    regions = {
        (row["district"].strip(), normalize_dong_name(row["dong"])): {
            "adm_cd": row["adm_cd"].strip(),
            "district": row["district"].strip(),
            "dong": row["dong"].strip(),
        }
        for row in rows
    }
    if len(regions) != 150:
        raise ValueError(f"기준 행정동은 150개여야 합니다: {len(regions)}개")
    return regions


def split_population_region(value: object) -> tuple[str, str] | None:
    """'대구광역시 구·군 행정동' 문자열에서 구·군과 행정동을 분리한다."""
    parts = str(value or "").strip().split()
    if len(parts) < 3 or parts[0] != "대구광역시":
        return None
    return parts[1], normalize_dong_name("".join(parts[2:]))


def read_and_aggregate_population(path: Path) -> dict[tuple[str, str], dict[str, int]]:
    """총인구·일반인·노인 인구를 읽고 출장소 인구를 본 행정동에 합산한다."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["인구현황"]
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=4, max_row=4))]
    try:
        region_index = header.index("행정구역")
        population_index = header.index("총 인구수")
        general_index = header.index("일반인 수")
        elderly_index = header.index("노인 수")
    except ValueError as error:
        raise ValueError("인구 파일에서 '행정구역', '총 인구수' 열을 찾지 못했습니다.") from error

    # 태경 commit merge 26-08-06 16-26
    totals: dict[tuple[str, str], dict[str, int]] = defaultdict(
        lambda: {"total": 0, "general": 0, "elderly": 0}
    )
    for row in worksheet.iter_rows(min_row=5, values_only=True):
        region = split_population_region(row[region_index])
        if region is None:
            continue
        population = row[population_index]
        if population in (None, ""):
            continue
        totals[region]["total"] += int(str(population).replace(",", ""))
        totals[region]["general"] += int(
            str(row[general_index] or 0).replace(",", "")
        )
        totals[region]["elderly"] += int(
            str(row[elderly_index] or 0).replace(",", "")
        )
    workbook.close()
    return dict(totals)


def correct_shelter_region(row: dict[str, str]) -> tuple[str, str, str | None]:
    """위치코드·주소로 확인한 군위군 산정면 표기와 불일치를 산성면으로 보정한다."""
    district = str(row.get("시군구") or "").strip()
    dong = normalize_dong_name(row.get("행정동"))
    reason = None
    is_sanseong = (
        str(row.get("위치코드") or "").strip() == "2772034000"
        or "군위군 산성면" in str(row.get("주소") or "")
    )
    if is_sanseong and (district, dong) != ("군위군", "산성면"):
        district, dong = "군위군", "산성면"
        reason = "산정면 표기·행정동 불일치의 산성면 보정"
    return district, dong, reason


def clean_text(value: object) -> str:
    """CSV의 결측 표현을 빈 문자열로 통일한다."""
    text = str(value or "").strip()
    return "" if text.lower() == "nan" else text


def build_schedule(row: dict[str, str], prefix: str) -> dict[str, object]:
    """기본·연장·추가 운영시간 열을 공통 일정 구조로 변환한다."""
    if prefix == "기본운영":
        enabled = True
    else:
        enabled = clean_text(row.get(f"{prefix}_여부")).upper() == "Y"
    return {
        "enabled": enabled,
        "days": clean_text(row.get(f"{prefix}_요일")),
        "start": clean_text(row.get(f"{prefix}_시작시간")),
        "end": clean_text(row.get(f"{prefix}_종료시간")),
    }


def preprocess(
    shelter_path: Path,
    population_path: Path,
    canonical_path: Path,
    output_dir: Path,
) -> tuple[Path, Path]:
    """두 원자료를 정규화하고 배포용 행정동 요약 CSV와 쉼터 JSON을 생성한다."""
    canonical = read_canonical_regions(canonical_path)
    population = read_and_aggregate_population(population_path)
    missing_population = sorted(set(canonical) - set(population))
    extra_population = sorted(set(population) - set(canonical))
    if missing_population or extra_population:
        raise ValueError(
            f"인구 행정동 결합 실패: 누락={missing_population}, 초과={extra_population}"
        )

    with shelter_path.open("r", encoding="utf-8-sig", newline="") as source:
        shelter_rows = list(csv.DictReader(source))

    details: list[dict[str, object]] = []
    summary: dict[str, dict[str, int]] = defaultdict(
        lambda: {"total": 0, "operating": 0, "public_operating": 0}
    )
    unmatched: list[str] = []
    correction_count = 0
    for row in shelter_rows:
        district, dong, correction = correct_shelter_region(row)
        region = canonical.get((district, dong))
        if region is None:
            unmatched.append(f"{row.get('No')}: {district} {dong}")
            continue
        correction_count += int(correction is not None)
        operating = clean_text(row.get("운영중여부")).upper() == "Y"
        public = clean_text(row.get("누구나이용가능여부")).upper() == "Y"
        adm_cd = region["adm_cd"]
        summary[adm_cd]["total"] += 1
        summary[adm_cd]["operating"] += int(operating)
        summary[adm_cd]["public_operating"] += int(operating and public)
        details.append(
            {
                "shelter_id": clean_text(row.get("No")),
                "adm_cd": adm_cd,
                "district": region["district"],
                "dong": region["dong"],
                "name": clean_text(row.get("쉼터명칭")),
                "operating": operating,
                "public_accessible": public,
                "capacity": clean_text(row.get("이용가능인원")),
                "schedules": [
                    build_schedule(row, "기본운영"),
                    build_schedule(row, "연장운영"),
                    build_schedule(row, "추가운영"),
                ],
                "source_location_code": clean_text(row.get("위치코드")),
                "correction": correction,
            }
        )
    if unmatched:
        raise ValueError(f"쉼터 행정동 결합 실패: {unmatched}")

    output_dir.mkdir(parents=True, exist_ok=True)
    summary_path = output_dir / "shelter_access_by_dong.csv"
    detail_path = output_dir / "shelter_access_detail.json"
    with summary_path.open("w", encoding="utf-8-sig", newline="") as target:
        fieldnames = [
            "adm_cd",
            "district",
            "dong",
            "population",
            "general_population",
            "elderly_population",
            "elderly_ratio_pct",
            "shelter_total",
            "shelter_operating",
            "shelter_public_operating",
            "public_operating_per_10000",
        ]
        writer = csv.DictWriter(target, fieldnames=fieldnames)
        writer.writeheader()
        for region in sorted(canonical.values(), key=lambda item: item["adm_cd"]):
            key = (region["district"], normalize_dong_name(region["dong"]))
            counts = summary[region["adm_cd"]]
            population_record = population[key]
            residents = population_record["total"]
            writer.writerow(
                {
                    **region,
                    "population": residents,
                    "general_population": population_record["general"],
                    "elderly_population": population_record["elderly"],
                    "elderly_ratio_pct": round(
                        population_record["elderly"] / residents * 100, 4
                    )
                    if residents
                    else 0,
                    "shelter_total": counts["total"],
                    "shelter_operating": counts["operating"],
                    "shelter_public_operating": counts["public_operating"],
                    "public_operating_per_10000": round(
                        counts["public_operating"] / residents * 10_000, 6
                    )
                    if residents
                    else 0,
                }
            )

    payload = {
        "schema_version": 1,
        "source": {
            "shelter": shelter_path.name,
            "population": population_path.name,
            "population_reference": "2026-07",
        },
        "preprocessing": {
            "branch_offices_aggregated": ["논공읍공단출장소", "다사읍서재출장소"],
            "sanseong_myeon_corrections": correction_count,
            "canonical_region_count": len(canonical),
            "shelter_count": len(details),
        },
        "shelters": details,
    }
    detail_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return summary_path, detail_path


def parse_args() -> argparse.Namespace:
    """명령행 입력 경로와 출력 경로를 읽는다."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--shelters", type=Path, required=True)
    parser.add_argument("--population", type=Path, required=True)
    parser.add_argument(
        "--canonical",
        type=Path,
        default=Path(__file__).resolve().parents[1]
        / "data"
        / "processed"
        / "heat_indicators.csv",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "data" / "processed",
    )
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_args()
    generated = preprocess(
        arguments.shelters,
        arguments.population,
        arguments.canonical,
        arguments.output_dir,
    )
    print("\n".join(str(path) for path in generated))
